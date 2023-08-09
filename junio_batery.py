

# -*- coding: utf-8 -*-
"""
Created on Sun Apr 26 16:50:58 2020

@author: ALEJANDRO
"""
from pyomo.environ import *
from pyomo.environ import *
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
def algoritmoJunio(cu,c,d,dicppv,dicpload, dicPrem, bateria):
    ini={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0}
    LimImp={1:200,2:200,3:203,4:210,5:210,6:213,7:73.32,8:0,9:0,10:0,11:0,12:0,13:0,14:0,15:0,16:0,17:261.07,18:558.82
    ,19:930,20:900,21:820,22:670,23:350,24:300}

    LimExp={1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:173.5,9:411.42,10:674.29,11:870.53,12:974.77,13:986.94,14:742.69,15:441.42,16:127.96,17:0,18:0,19:0,20:0,21:0,22:0,23:0,24:0}

    model = ConcreteModel()

    #definicion de variables

    model.tempo= Set(initialize=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24])

    model.Pload=Param(model.tempo, initialize=dicpload)

    model.Ppv=Param(model.tempo, initialize=dicppv)








    model.Prem=Param(model.tempo, initialize=dicPrem)

    model.Expt=Var(model.tempo)
    model.Pgrid=Var(model.tempo) 

    def f(model, i):
        return (ini[i], LimExp[i])
    model.Exp2 = Var(model.tempo, bounds=f,doc=' Var Exportacion_2')


    def f(model, i):
        return (ini[i], LimImp[i])
    model.Exp1 = Var(model.tempo, bounds=f,doc=' Var Exportacion_1')

    def f(model, i):
        return (ini[i], LimImp[i])
    model.Imp = Var(model.tempo, bounds=f, doc=' Var Importacion')
    #------------------------------------------------------------------------------
    #variables de la bateria plomo acido GAMS

    SOC0=800
    Socmax=3420
    ata_c=0.95
    eta_d=0.9

    model.Socbat=Var(model.tempo, bounds=(684,3420), doc=' Socmin y Socmax Bateria')
    model.Pcbat=Var(model.tempo,bounds=(0,3078), doc=' Pcarga bateria' )
    model.Pdbat=Var(model.tempo,bounds=(0,3078), doc=' Pdescarga bateria' )


    #variables binbaria  de la bateria plomo acido
    model.Idbat=Var(model.tempo, within=Binary, doc=' variable descarga bateria plomo acido')
    model.Icbat=Var(model.tempo, within=Binary, doc=' variable carga bateria plomo acido')
    #------------------------------------------------------------------------------


    #   ecuacion balance(t)..        Prem(t)=e=Pgrid(t)+Pc(t)-Pd(t);  
    def c_init0(model,i):
        return model.Expt[i]==model.Exp2[i]+model.Exp1[i]
    model.cond0= Constraint(model.tempo,rule=c_init0)

    def c_init01(model,i):
        return model.Pgrid[i]==model.Expt[i]-model.Imp[i]
    model.cond01= Constraint(model.tempo,rule=c_init01,doc=' Pgrid')

    def c_init02(model,i):
        return model.Exp1[i]<=model.Exp2[i]
    model.cond02= Constraint(model.tempo,rule=c_init02,doc=' Pgrid')

    if(bateria=="si"):
        def c_init1(model,i):
            return model.Prem[i]==model.Pgrid[i]+model.Pcbat[i]-model.Pdbat[i]
        model.cond1= Constraint(model.tempo,rule=c_init1)

        #restriciones de la bateria plomo acido

        #ecuacion3(t)..      Pc(t)=l= 0.9*SOCmax*Icba(t);
        def c_init2(model,i):
            return model.Pcbat[i]<=0.9*Socmax*model.Icbat[i]
        model.cond2= Constraint(model.tempo,rule=c_init2)


        #ecuacion4(t)..      Pd(t)=l= 0.9*SOCmax*Idba(t);
        def c_init3(model,i):
            return model.Pdbat[i]<=0.9*Socmax*model.Idbat[i]
        model.cond3= Constraint(model.tempo,rule=c_init3)


        #ecuacion5(t)..      (1-Icba(t))+(1-Idba(t))=e=1;
        def c_init4(model,i):
            return   (1-model.Icbat[i])+(1-model.Idbat[i])==1
            # return model.Icbat[i]+model.Idbat[i]<=1
        model.cond4= Constraint(model.tempo,rule=c_init4)
        #------------------------------------------------------------------------------
        #restriciones SOC bateria

        model.cond5= Constraint(expr=model.Socbat[1]==SOC0+(model.Pcbat[1]*ata_c)-(model.Pdbat[1]/eta_d))

        model.cond6= Constraint(expr=model.Socbat[2]==model.Socbat[1]+(model.Pcbat[2]*ata_c)-(model.Pdbat[2]/eta_d))

        model.cond7= Constraint(expr=model.Socbat[3]==model.Socbat[2]+(model.Pcbat[3]*ata_c)-(model.Pdbat[3]/eta_d))

        model.cond8= Constraint(expr=model.Socbat[4]==model.Socbat[3]+(model.Pcbat[4]*ata_c)-(model.Pdbat[4]/eta_d))

        model.cond9= Constraint(expr=model.Socbat[5]==model.Socbat[4]+(model.Pcbat[5]*ata_c)-(model.Pdbat[5]/eta_d))

        model.cond10= Constraint(expr=model.Socbat[6]==model.Socbat[5]+(model.Pcbat[6]*ata_c)-(model.Pdbat[6]/eta_d))

        model.cond11= Constraint(expr=model.Socbat[7]==model.Socbat[6]+(model.Pcbat[7]*ata_c)-(model.Pdbat[7]/eta_d))

        model.cond12= Constraint(expr=model.Socbat[8]==model.Socbat[7]+(model.Pcbat[8]*ata_c)-(model.Pdbat[8]/eta_d))

        model.cond13= Constraint(expr=model.Socbat[9]==model.Socbat[8]+(model.Pcbat[9]*ata_c)-(model.Pdbat[9]/eta_d))
        
        model.cond14= Constraint(expr=model.Socbat[10]==model.Socbat[9]+(model.Pcbat[10]*ata_c)-(model.Pdbat[10]/eta_d))

        model.cond15= Constraint(expr=model.Socbat[11]==model.Socbat[10]+(model.Pcbat[11]*ata_c)-(model.Pdbat[11]/eta_d))
        
        model.cond16= Constraint(expr=model.Socbat[12]==model.Socbat[11]+(model.Pcbat[12]*ata_c)-(model.Pdbat[12]/eta_d))

        model.cond17= Constraint(expr=model.Socbat[13]==model.Socbat[12]+(model.Pcbat[13]*ata_c)-(model.Pdbat[13]/eta_d))

        model.cond18= Constraint(expr=model.Socbat[14]==model.Socbat[13]+(model.Pcbat[14]*ata_c)-(model.Pdbat[14]/eta_d))

        model.cond19= Constraint(expr=model.Socbat[15]==model.Socbat[14]+(model.Pcbat[15]*ata_c)-(model.Pdbat[15]/eta_d))

        model.cond20= Constraint(expr=model.Socbat[16]==model.Socbat[15]+(model.Pcbat[16]*ata_c)-(model.Pdbat[16]/eta_d))

        model.cond21= Constraint(expr=model.Socbat[17]==model.Socbat[16]+(model.Pcbat[17]*ata_c)-(model.Pdbat[17]/eta_d))

        model.cond22= Constraint(expr=model.Socbat[18]==model.Socbat[17]+(model.Pcbat[18]*ata_c)-(model.Pdbat[18]/eta_d))

        model.cond23= Constraint(expr= model.Socbat[19]==model.Socbat[18]+(model.Pcbat[19]*ata_c)-(model.Pdbat[19]/eta_d))     

        model.cond24= Constraint(expr= model.Socbat[20]==model.Socbat[19]+(model.Pcbat[20]*ata_c)-(model.Pdbat[20]/eta_d))  

        model.cond25= Constraint(expr= model.Socbat[21]==model.Socbat[20]+(model.Pcbat[21]*ata_c)-(model.Pdbat[21]/eta_d))  

        model.cond26= Constraint(expr= model.Socbat[22]==model.Socbat[21]+(model.Pcbat[22]*ata_c)-(model.Pdbat[22]/eta_d))  

        model.cond27= Constraint(expr= model.Socbat[23]==model.Socbat[22]+(model.Pcbat[23]*ata_c)-(model.Pdbat[23]/eta_d))  

        model.cond28= Constraint(expr= model.Socbat[24]==model.Socbat[23]+(model.Pcbat[24]*ata_c)-(model.Pdbat[24]/eta_d))

        model.cond29= Constraint(expr= model.Socbat[24]==800)
        #-----------------------------------------------------------------------------

        # funcion objetivo
        #scalar  Cu/0.405/,C /0.063/,D /0.142/;   
        Cu=cu
        C=c
        D=d
        model.value = Objective(expr = sum( (Socmax-model.Socbat[i])*0.1*Cu for i in model.tempo)+sum((model.Exp1[i]-model.Imp[i])*Cu-(model.Exp1[i]*C)+(model.Exp2[i]*D) for i in model.tempo),sense = minimize )
        #-----------------------------------------------------------------------------

        def pyomo_postprocess(options=None, instance=None, results=None):
            model.Prem.display()


        opt = SolverFactory('glpk', executable='C:/Users/USUARIO/.conda/envs/glpk-env/Library/bin/glpsol.exe')
        result_obj = opt.solve(model, tee=True)
        model.display()

         # Obtener los valores de Exp1
        exp1_values = [value for _, value in model.Exp1.get_values().items()]

        # Obtener los valores de Exp2
        exp2_values = [value for _, value in model.Exp2.get_values().items()]

        # Obtener los valores de Imp
        imp_values = [value for _, value in model.Imp.get_values().items()]

        # Crear una lista de tiempo para el eje x (asumiendo que hay 24 periodos)
        tiempo = range(1, 25)

    

        # Obtener los valores de Exp1, Exp2 e Imp
        exp1_values = [value(model.Exp1[i]) for i in model.tempo]
        exp2_values = [value(model.Exp2[i]) for i in model.tempo]
        imp_values = [value(model.Imp[i]) for i in model.tempo]

        # Crear un DataFrame de pandas con los valores obtenidos
        data = {'Exp1': exp1_values, 'Exp2': exp2_values, 'Imp': imp_values}
        df = pd.DataFrame(data)

        excel_filename='Calculo_Mes_Junio_Bateria.xlsx'


        # Guardar el DataFrame en un archivo de Excel
        df.to_excel(excel_filename, index=False)

        # Leer el archivo de Excel
        wb = Workbook()
        wb = openpyxl.load_workbook(excel_filename)
        sheet = wb.active

        # Graficar los resultados
        plt.plot(tiempo, exp1_values, label='Exp1')
        plt.plot(tiempo, exp2_values, label='Exp2')
        plt.plot(tiempo, imp_values, label='Imp')
        plt.xlabel('Tiempo')
        plt.ylabel('Valor')
        plt.legend()

        graph_filename = 'graph_Junio_Bateria.png'
        plt.savefig(graph_filename)
        plt.close()

        # Insertar la imagen de la gráfica en el archivo de Excel
        img = Image(graph_filename)
        img.anchor = 'E2'  # Cambia la celda de anclaje según tus necesidades
        sheet.add_image(img)

        # Guardar el archivo de Excel
        wb.save(excel_filename)
        wb.close()
    if(bateria=="no"): 
        def c_init1(model,i):
            return model.Prem[i]==model.Pgrid[i]
        model.cond1= Constraint(model.tempo,rule=c_init1)
        
        #ecuacion3(t)..      Pc(t)=l= 0.9*SOCmax*Icba(t);
       
        model.cond2= Constraint(model.tempo)


        #ecuacion4(t)..      Pd(t)=l= 0.9*SOCmax*Idba(t);
        model.cond3= Constraint(model.tempo)


        #ecuacion5(t)..      (1-Icba(t))+(1-Idba(t))=e=1;
        model.cond4= Constraint(model.tempo)
        #------------------------------------------------------------------------------
        # funcion objetivo

        Cu=cu
        C=c
        D=d
        model.value = Objective(expr = sum( 0.1*Cu for i in model.tempo)+sum((model.Exp1[i]-model.Imp[i])*Cu-(model.Exp1[i]*C)+(model.Exp2[i]*D) for i in model.tempo),sense = minimize )

    #-----------------------------------------------------------------------------
        def pyomo_postprocess(options=None, instance=None, results=None):
            model.Prem.display()

        #----------------------------------------------------------------------------------------------------------------------------------
        opt = SolverFactory('glpk',  executable='C:/Users/USUARIO/.conda/envs/glpk-env/Library/bin/glpsol.exe')
        result_obj = opt.solve(model)
        model.display()

        # Obtener los valores de Exp1
        exp1_values = [value for _, value in model.Exp1.get_values().items()]

        # Obtener los valores de Exp2
        exp2_values = [value for _, value in model.Exp2.get_values().items()]

        # Obtener los valores de Imp
        imp_values = [value for _, value in model.Imp.get_values().items()]

        # Crear una lista de tiempo para el eje x (asumiendo que hay 24 periodos)
        tiempo = range(1, 25)

    

        # Obtener los valores de Exp1, Exp2 e Imp
        exp1_values = [value(model.Exp1[i]) for i in model.tempo]
        exp2_values = [value(model.Exp2[i]) for i in model.tempo]
        imp_values = [value(model.Imp[i]) for i in model.tempo]

        # Crear un DataFrame de pandas con los valores obtenidos
        data = {'Exp1': exp1_values, 'Exp2': exp2_values, 'Imp': imp_values}
        df = pd.DataFrame(data)

        excel_filename='Calculo_Mes_Junio_No_Bateria.xlsx'


        # Guardar el DataFrame en un archivo de Excel
        df.to_excel(excel_filename, index=False)

        # Leer el archivo de Excel
        wb = Workbook()
        wb = openpyxl.load_workbook(excel_filename)
        sheet = wb.active

        # Graficar los resultados
        plt.plot(tiempo, exp1_values, label='Exp1')
        plt.plot(tiempo, exp2_values, label='Exp2')
        plt.plot(tiempo, imp_values, label='Imp')
        plt.xlabel('Tiempo')
        plt.ylabel('Valor')
        plt.legend()

        graph_filename = 'graph_Junio_No_Bateria.png'
        plt.savefig(graph_filename)
        plt.close()

        # Insertar la imagen de la gráfica en el archivo de Excel
        img = Image(graph_filename)
        img.anchor = 'E2'  # Cambia la celda de anclaje según tus necesidades
        sheet.add_image(img)

        # Guardar el archivo de Excel
        wb.save(excel_filename)
        wb.close()



